#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
把 extract_pdf.py 辨識出來的「他機關」資料，併入 convert.py 產生的 CGSS 輸出檔。

用法：
    python merge_agencies.py CGSS_轉換結果.xlsx output/all_agencies.json final.xlsx

比對方式：用 CGSS 輸出檔裡的 CustomID（也就是 TEST.xlsx 的「序號」）
跟 all_agencies.json 裡的 serial 對應。

⚠️ 多筆機關名稱/金額目前用「；」分號組成同一個儲存格
   （UnitName / ExtAmount 欄位本身標註「可多筆」，但 CGSS 系統實際規定的
   多筆分隔方式建議上傳前先跟系統管理端確認一次，此處分號是暫時的合理假設）。
"""

import sys
import json
from openpyxl import load_workbook

C_CUSTOM_ID = 1
C_UNIT_NAME = 19
C_EXT_AMOUNT = 20


def main():
    if len(sys.argv) != 4:
        print("用法: python merge_agencies.py CGSS_轉換結果.xlsx all_agencies.json final.xlsx")
        sys.exit(1)

    cgss_path, json_path, out_path = sys.argv[1], sys.argv[2], sys.argv[3]

    with open(json_path, "r", encoding="utf-8") as f:
        records = json.load(f)

    by_serial = {r["serial"]: r for r in records if r.get("serial")}

    wb = load_workbook(cgss_path)
    ws = wb.active

    matched, skipped = 0, []
    for row in range(11, ws.max_row + 1):
        custom_id = ws.cell(row=row, column=C_CUSTOM_ID).value
        if not custom_id:
            continue
        rec = by_serial.get(str(custom_id).strip())
        if not rec or rec.get("_error") or rec.get("_parse_error"):
            skipped.append(custom_id)
            continue
        agencies = rec.get("other_agencies") or []
        if not agencies:
            continue
        names = "；".join(a.get("name", "") for a in agencies if a.get("name"))
        amounts = "；".join(str(a.get("amount", "")) for a in agencies if a.get("amount"))
        ws.cell(row=row, column=C_UNIT_NAME, value=names)
        ws.cell(row=row, column=C_EXT_AMOUNT, value=amounts)
        matched += 1

    wb.save(out_path)
    print(f"已合併 {matched} 筆他機關資料到 {out_path}")
    if skipped:
        print(f"以下序號沒有對應的辨識結果或辨識失敗，UnitName/ExtAmount 留白：{skipped}")


if __name__ == "__main__":
    main()
