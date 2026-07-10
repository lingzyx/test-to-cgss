#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
TEST.xlsx (自建睦鄰案件資料) → CGSS 批次上傳格式 轉換工具
=========================================================

用法：
    python convert.py TEST.xlsx CGSS.xlsx output.xlsx

參數：
    TEST.xlsx   自建的睦鄰案件資料（來源檔）
    CGSS.xlsx   CGSS 批次上傳範例／模板檔（決定輸出欄位與版型）
    output.xlsx 產出的 CGSS 格式檔案

這支程式把「來源欄位可以百分之百對應」的欄位自動填入，
無法從 TEST.xlsx 判斷出來的必填欄位（例如 ApplyDate 申請日期）
一律留白，並在轉換結束後印出清單，提醒需要人工補齊的資料列。

===========================================================
★ 這份對照表是根據 CGSS.xlsx 裡「已經存在的 2 筆真實範例列」
   (第11、12列) 反推比對 TEST.xlsx 對應資料列 (第19、20列) 得出，
   而不是憑空假設，所以下列常數欄位的預設值有實際依據：
     Fund_Category (經費類別)      = 公務預算
     Fund_Name (工作/業務計畫名稱) = U200教育文化
     Organization_Type (團體類型)  = 扣繳編號(統一編號)
     CountyName (所屬地區)         = 南投縣
     IsPublic (開放查詢註記)       = 是
   如果換年度／換計畫名稱，請直接修改下面 CONFIG 區塊即可，
   不需要改動其他程式碼。
===========================================================
"""

import sys
import re
import datetime
from openpyxl import load_workbook

# ============================================================
# CONFIG — 每年 / 每個計畫只需要改這裡
# ============================================================
CONFIG = {
    "IS_PUBLIC": "是",                     # *開放查詢註記
    "FUND_CATEGORY": "公務預算",            # *經費類別
    "FUND_NAME": "U200教育文化",            # *工作/業務計畫名稱
    "ORGANIZATION_TYPE": "扣繳編號(統一編號)",  # *民間團體類型
    "COUNTY_NAME": "南投縣",                # *民間團體所歸屬之地區
    "FIRST_DATA_ROW": 11,                  # CGSS 範本資料從第幾列開始寫（找不到就用這個預設值）
    "ONLY_ROWS_MARKED_CGSS": True,          # 是否只轉換 TEST 裡「CGSS」欄有標記 v 的列
}

# ------------------------------------------------------------
# 欄位定位改用「欄名比對」，不用寫死的欄位編號。
# 這樣就算 TEST.xlsx 或 CGSS.xlsx 的欄位順序被調整過，
# 只要欄名文字還在，程式一樣找得到對應的資料，不會跑位。
# 找不到必要欄位時會直接中止並列出缺少哪些欄，而不是默默把資料塞錯欄。
# ------------------------------------------------------------

def _norm(s):
    """欄名正規化：去除換行、全形/半形空白，方便比對。"""
    if s is None:
        return ""
    return re.sub(r"[\s\u3000]+", "", str(s))


# TEST.xlsx 欄位比對規則：(欄位代號, 必須包含的關鍵字, 不能包含的關鍵字, 是否必要)
TEST_FIELD_RULES = [
    ("SERIAL",        ["序號"],           [],         True),
    ("ORG_NAME",      ["申請單位"],        [],         True),
    ("ORG_ID",        ["統一編號"],        [],         True),
    ("ACTIVITY_NAME", ["活動名稱"],        [],         True),
    ("START",         ["活動時間", "起"],  [],         True),
    ("END",           ["活動時間", "迄"],  [],         True),
    ("SELF_FUND",     ["自有款"],          [],         True),
    ("APPLIED_AMT",   ["補助金額"],        ["核定", "他機關"], True),   # 申請補助金額
    ("CONFIRMED_AMT", ["補助金額"],        ["申請", "他機關"], True),   # 核定補助金額
    ("CONFIRM_DATE",  ["核定日期"],        [],         True),
    ("CGSS_FLAG",     ["CGSS"],           ["睦鄰"],    True),
]


def build_test_column_map(ws, header_row=1):
    """在 TEST.xlsx 的表頭列裡，用欄名關鍵字找出每個欄位對應的欄號。"""
    max_col = ws.max_column
    headers = [_norm(ws.cell(row=header_row, column=c).value) for c in range(1, max_col + 1)]

    col_map = {}
    missing = []
    for field, must_have, must_not_have, required in TEST_FIELD_RULES:
        found_col = None
        for idx, h in enumerate(headers, start=1):
            if not h:
                continue
            if all(k in h for k in must_have) and not any(k in h for k in must_not_have):
                found_col = idx
                break
        if found_col is None:
            if required:
                missing.append(field)
        else:
            col_map[field] = found_col

    if missing:
        raise ValueError(
            "TEST.xlsx 第 " + str(header_row) + " 列表頭裡找不到以下必要欄位，"
            "請確認欄位名稱是否被更動過：" + "、".join(missing)
        )
    return col_map


# CGSS.xlsx 欄位比對：直接用英文欄名（第 5 列，例如 CustomID、Organization_Name）
CGSS_REQUIRED_FIELDS = [
    "CustomID", "IsPublic", "ApplyDate", "Fund_Category", "Fund_Name", "Fund_Year",
    "Organization_Name", "Organization_ID", "Organization_Type", "CountyName",
    "Subject", "StartDate", "EndDate", "SelfProvidedAmount", "Applied",
    "ConfirmDate", "ConfirmFund_Name", "ConfirmFund_Year", "ConfirmedAmount",
]


def build_cgss_column_map(ws, max_scan_rows=15):
    """掃描 CGSS.xlsx 前面幾列，找出英文欄名所在的那一列，
    並回傳 {欄名: 欄號} 對照表，以及那一列是第幾列（資料就接在它下面）。"""
    header_row = None
    headers = None
    for r in range(1, max_scan_rows + 1):
        row_vals = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
        if "Organization_Name" in row_vals:
            header_row = r
            headers = row_vals
            break

    if header_row is None:
        raise ValueError("在 CGSS.xlsx 前 " + str(max_scan_rows) + " 列裡找不到英文欄名列（例如 Organization_Name），"
                          "請確認這是不是正確的 CGSS 範本檔。")

    col_map = {}
    for idx, h in enumerate(headers, start=1):
        if h:
            col_map[str(h).strip()] = idx

    missing = [f for f in CGSS_REQUIRED_FIELDS if f not in col_map]
    if missing:
        raise ValueError("CGSS.xlsx 表頭裡找不到以下必要欄位，請確認範本是否為官方最新版本：" + "、".join(missing))

    return col_map, header_row


# 這兩個欄位不是必要欄位，但如果找得到就會使用（找不到就靜靜跳過，不影響轉換）
CGSS_OPTIONAL_FIELDS = ["UnitName", "ExtAmount"]


def parse_roc_date(value):
    """把 TEST.xlsx 裡各種格式的民國日期轉成 datetime.date。
    支援: 115/2/26、115.5.04、115. 5. 19、115年3月18日 13:00-17:00、
    115/04/12(日) 08:30~12:20、已經是 datetime 物件 的情況。
    無法解析時回傳 None。"""
    if value is None:
        return None
    if isinstance(value, datetime.datetime):
        return value.date()
    if isinstance(value, datetime.date):
        return value
    s = str(value).strip()
    if not s:
        return None
    m = re.search(r"(\d{2,3})\s*[/.年]\s*(\d{1,2})\s*[/.月]\s*(\d{1,2})", s)
    if not m:
        return None
    roc_year, month, day = int(m.group(1)), int(m.group(2)), int(m.group(3))
    try:
        return datetime.date(roc_year + 1911, month, day)
    except ValueError:
        return None


def normalize_org_id(value):
    """統一編號一律轉成去除空白、補滿 8 碼的文字字串。"""
    if value is None:
        return ""
    s = str(value).strip()
    if s.isdigit():
        s = s.zfill(8)
    return s


def convert(test_path, cgss_path, output_path):
    warnings = []

    test_wb = load_workbook(test_path, data_only=True)
    test_ws = test_wb.active
    T = build_test_column_map(test_ws)

    rows_out = []
    r = 2
    while True:
        serial = test_ws.cell(row=r, column=T["SERIAL"]).value
        if serial is None:
            # 允許中間偶有空列，但連續 3 列都空就視為資料結束
            blank_ahead = all(
                test_ws.cell(row=r + i, column=T["SERIAL"]).value is None
                for i in range(3)
            )
            if blank_ahead:
                break
            r += 1
            continue

        flag = test_ws.cell(row=r, column=T["CGSS_FLAG"]).value
        flagged = str(flag).strip().lower() == "v" if flag else False
        if CONFIG["ONLY_ROWS_MARKED_CGSS"] and not flagged:
            warnings.append(f"第{r}列（序號 {serial}）CGSS 欄未標記 v，已略過未輸出。")
            r += 1
            continue

        fund_year_match = re.match(r"^(\d+)-", str(serial).strip())
        fund_year = fund_year_match.group(1) if fund_year_match else ""

        start_date = parse_roc_date(test_ws.cell(row=r, column=T["START"]).value)
        end_date = parse_roc_date(test_ws.cell(row=r, column=T["END"]).value)
        if end_date is None and start_date is not None:
            warnings.append(
                f"第{r}列（序號 {serial}）活動時間(迄) 無法解析或含多個日期，"
                f"已暫用活動時間(起) {start_date} 作為結束日，請人工確認。"
            )
            end_date = start_date
        if start_date is None:
            warnings.append(f"第{r}列（序號 {serial}）活動時間(起) 無法解析，StartDate 留白。")

        confirm_date = parse_roc_date(test_ws.cell(row=r, column=T["CONFIRM_DATE"]).value)
        confirmed_amt = test_ws.cell(row=r, column=T["CONFIRMED_AMT"]).value
        if confirmed_amt not in (None, "") and confirm_date is None:
            warnings.append(
                f"第{r}列（序號 {serial}）已有核定補助金額但無核定日期，ConfirmDate 留白，請人工補齊。"
            )

        warnings.append(
            f"第{r}列（序號 {serial}）ApplyDate（*申請日期，必填）在 TEST.xlsx 中無對應欄位，已留白，請人工補齊。"
        )

        rows_out.append({
            "custom_id": str(serial).strip(),
            "apply_date": None,  # 無法從來源判斷，見上方 warnings
            "fund_category": CONFIG["FUND_CATEGORY"],
            "fund_name": CONFIG["FUND_NAME"],
            "fund_year": fund_year,
            "org_name": str(test_ws.cell(row=r, column=T["ORG_NAME"]).value or "").strip(),
            "org_id": normalize_org_id(test_ws.cell(row=r, column=T["ORG_ID"]).value),
            "org_type": CONFIG["ORGANIZATION_TYPE"],
            "county_name": CONFIG["COUNTY_NAME"],
            "subject": str(test_ws.cell(row=r, column=T["ACTIVITY_NAME"]).value or "").strip(),
            "start_date": start_date,
            "end_date": end_date,
            "self_amount": test_ws.cell(row=r, column=T["SELF_FUND"]).value,
            "applied_amount": test_ws.cell(row=r, column=T["APPLIED_AMT"]).value,
            "confirm_date": confirm_date,
            "confirmed_amount": confirmed_amt,
        })
        r += 1

    cgss_wb = load_workbook(cgss_path)
    cgss_ws = cgss_wb.active
    C, header_row = build_cgss_column_map(cgss_ws)
    first = header_row + 1  # 資料緊接在英文欄名列的下一列

    # 清掉範本裡原本示範用的資料列，避免與新資料混在一起
    max_existing_row = cgss_ws.max_row
    if max_existing_row >= first:
        cgss_ws.delete_rows(first, max_existing_row - first + 1)

    for i, row in enumerate(rows_out):
        rr = first + i
        cgss_ws.cell(row=rr, column=C["CustomID"], value=row["custom_id"])
        cgss_ws.cell(row=rr, column=C["IsPublic"], value=CONFIG["IS_PUBLIC"])
        cgss_ws.cell(row=rr, column=C["ApplyDate"], value=row["apply_date"])
        cgss_ws.cell(row=rr, column=C["Fund_Category"], value=row["fund_category"])
        cgss_ws.cell(row=rr, column=C["Fund_Name"], value=row["fund_name"])
        cgss_ws.cell(row=rr, column=C["Fund_Year"], value=row["fund_year"])
        cgss_ws.cell(row=rr, column=C["Organization_Name"], value=row["org_name"])
        cgss_ws.cell(row=rr, column=C["Organization_ID"], value=row["org_id"])
        cgss_ws.cell(row=rr, column=C["Organization_Type"], value=row["org_type"])
        cgss_ws.cell(row=rr, column=C["CountyName"], value=row["county_name"])
        cgss_ws.cell(row=rr, column=C["Subject"], value=row["subject"])
        cgss_ws.cell(row=rr, column=C["StartDate"], value=row["start_date"])
        cgss_ws.cell(row=rr, column=C["EndDate"], value=row["end_date"])
        cgss_ws.cell(row=rr, column=C["SelfProvidedAmount"], value=str(row["self_amount"]) if row["self_amount"] is not None else "")
        cgss_ws.cell(row=rr, column=C["Applied"], value=str(row["applied_amount"]) if row["applied_amount"] is not None else "")
        if row["confirm_date"] is not None:
            cgss_ws.cell(row=rr, column=C["ConfirmDate"], value=row["confirm_date"].strftime("%Y-%m-%d"))
            cgss_ws.cell(row=rr, column=C["ConfirmFund_Name"], value=row["fund_name"])
            cgss_ws.cell(row=rr, column=C["ConfirmFund_Year"], value=row["fund_year"])
        if row["confirmed_amount"] not in (None, ""):
            cgss_ws.cell(row=rr, column=C["ConfirmedAmount"], value=row["confirmed_amount"])

    cgss_wb.save(output_path)

    print(f"轉換完成：共輸出 {len(rows_out)} 筆資料到 {output_path}")
    print(f"\n共有 {len(warnings)} 項提醒，請於上傳 CGSS 前逐一確認：\n")
    for w in warnings:
        print(" -", w)

    return rows_out, warnings


if __name__ == "__main__":
    if len(sys.argv) != 4:
        print("用法: python convert.py TEST.xlsx CGSS.xlsx output.xlsx")
        sys.exit(1)
    try:
        convert(sys.argv[1], sys.argv[2], sys.argv[3])
    except ValueError as e:
        print("轉換中止：", e)
        sys.exit(1)
